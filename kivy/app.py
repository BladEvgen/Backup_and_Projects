import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from datetime import datetime, timedelta
from openpyxl import load_workbook
from kivy.core.window import Window
import os

kivy.require('1.9.0')  # Minimum Kivy version


class ScheduleApp(App):
    def build(self):
        # Main layout
        layout = BoxLayout(orientation="vertical", spacing=10, padding=10)

        # Get the current date and date range
        end_date = datetime.now() + timedelta(days=12)

        # Header label with the date range
        date_range_label = Label(
            text=f'Расписание на {end_date.strftime("%d.%m.%Y")}',
            color=(0, 0, 1, 1),
            size_hint=(1, None),
            height=50,
            font_size="24sp",
        )
        layout.add_widget(date_range_label)

        try:
            xlsx_file_path = self.locate_schedule_xlsx()
            if xlsx_file_path:
                wb = load_workbook(xlsx_file_path)
                sheet = wb.active

                # Create a BoxLayout to display the schedule
                schedule_layout = BoxLayout(
                    orientation="vertical", spacing=10, padding=10
                )
                data_found = False

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_date = row[0]
                    row_time = row[1]
                    row_discipline = row[2]
                    row_aud = row[3]
                    row_teacher = row[7]
                    row_column_G = row[6]  # Column G

                    if (
                        isinstance(row_date, datetime)
                        and row_date.date() == end_date.date()
                        and (
                            "__________" in row_column_G
                            or "2 подгруппа" in row_column_G
                        )
                    ):
                        data_found = True

                        # Grid layout for each row of data
                        row_layout = GridLayout(cols=2, spacing=5)
                        row_layout.add_widget(Label(text="Время:", font_size="18sp"))
                        row_layout.add_widget(
                            Label(text=str(row_time), font_size="18sp")
                        )

                        # Wrap "Discipline" text onto a new line if it exceeds 30 characters
                        discipline_text = str(row_discipline)
                        if len(discipline_text) > 32:
                            wrapped_discipline = "\n                        ".join(
                                [
                                    discipline_text[i : i + 32]
                                    for i in range(0, len(discipline_text), 32)
                                ]
                            )
                            row_layout.add_widget(
                                Label(text="Предмет:", font_size="18sp")
                            )
                            row_layout.add_widget(
                                Label(text=wrapped_discipline, font_size="18sp")
                            )
                        else:
                            row_layout.add_widget(
                                Label(text="Предмет:", font_size="18sp")
                            )
                            row_layout.add_widget(
                                Label(text=discipline_text, font_size="18sp")
                            )

                        row_layout.add_widget(
                            Label(text="Аудитория:", font_size="18sp")
                        )
                        row_layout.add_widget(
                            Label(
                                text=str(row_aud if row_aud else "N/A"),
                                font_size="18sp",
                            )
                        )
                        row_layout.add_widget(
                            Label(text="Преподаватель:", font_size="18sp")
                        )
                        row_layout.add_widget(
                            Label(text=str(row_teacher), font_size="18sp")
                        )

                        schedule_layout.add_widget(row_layout)

                if not data_found:
                    schedule_layout.add_widget(
                        Label(
                            text="[b]На сегодня ничего не было найдено[/b]",
                            size_hint=(1, None),
                            height=350,
                            markup=True,
                            font_size="35sp",
                        )
                    )

                # Scrollview for the schedule
                scrollview = ScrollView(size_hint=(1, 1))
                scrollview.add_widget(schedule_layout)
                layout.add_widget(scrollview)

        except Exception as e:
            error_popup = Popup(
                title="Error",
                content=Label(text=f"Ошибка: {str(e)}", font_size="18sp"),
                size_hint=(None, None),
                size=(400, 200),
            )
            error_popup.open()
            print(str(e))

        return layout

    def locate_schedule_xlsx(self):
        documents_dir = os.path.join(os.path.expanduser("~"), "Documents")
        xlsx_path = os.path.join(documents_dir, "schedule.xlsx")
        if os.path.exists(xlsx_path):
            return xlsx_path
        return None


if __name__ == "__main__":
    ScheduleApp().run()
