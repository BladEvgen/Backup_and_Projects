import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from datetime import datetime, timedelta
from openpyxl import load_workbook
from kivy.core.window import Window
import os


class ScheduleApp(App):
    def build(self):
        layout = BoxLayout(orientation="vertical", spacing=10, padding=10)

        # Get the current date and date range
        end_date = datetime.now() + timedelta(days=6)

        # Add a label for the date range
        date_range_label = Label(
            text=f'Schedule for {end_date.strftime("%d.%m.%Y")}',
            color=(1, 0.5, 0, 1),
            size_hint=(1, None),
            height=50,
        )
        layout.add_widget(date_range_label)

        try:
            xlsx_file_path = self.locate_schedule_xlsx()
            if xlsx_file_path:
                wb = load_workbook(xlsx_file_path)
                sheet = wb.active

                # Create a BoxLayout to display the schedule as a continuous block
                schedule_layout = BoxLayout(
                    orientation="vertical", spacing=10, padding=10
                )

                data_found = False

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_date = row[0]
                    row_time = row[1]
                    row_discipline = row[2]
                    row_aud = row[3]
                    row_teacher = row[7]
                    row_column_G = row[6]  # Column G

                    # Check if the row date falls within the specified date range
                    if (
                        isinstance(row_date, datetime)
                        and row_date.date() == end_date.date()
                        and (
                            "__________" in row_column_G
                            or "2 подгруппа" in row_column_G
                        )
                    ):
                        data_found = True

                        # Create a BoxLayout for each row of data
                        row_layout = GridLayout(cols=2, spacing=5)
                        row_layout.add_widget(Label(text="Time:"))
                        row_layout.add_widget(Label(text=str(row_time)))
                        row_layout.add_widget(Label(text="Discipline:"))
                        row_layout.add_widget(Label(text=str(row_discipline)))
                        row_layout.add_widget(Label(text="Audience:"))
                        row_layout.add_widget(
                            Label(text=str(row_aud if row_aud else "N/A"))
                        )
                        row_layout.add_widget(Label(text="Teacher:"))
                        row_layout.add_widget(Label(text=str(row_teacher)))

                        schedule_layout.add_widget(row_layout)

                if not data_found:
                    schedule_layout.add_widget(
                        Label(
                            text="[b]Nothing found for today[/b]",
                            size_hint=(1, None),
                            height=350,
                            markup=True,
                            font_size="35sp",
                        )
                    )

                scrollview = ScrollView(
                    size_hint=(1, None), size=(Window.width, Window.height - 200)
                )
                scrollview.add_widget(schedule_layout)
                layout.add_widget(scrollview)

        except Exception as e:
            error_popup = Popup(
                title="Error",
                content=Label(text=f"An error occurred: {str(e)}"),
                size_hint=(None, None),
                size=(400, 200),
            )
            error_popup.open()
            print(str(e))
        return layout

    def locate_schedule_xlsx(self):
        documents_dir = os.path.join(os.path.expanduser("~"), "Documents")
        xlsx_path = os.path.join(documents_dir, "schedule.xlsx")
        if os.path.exists(xlsx_path):
            return xlsx_path
        return None

    def load_xlsx(self, _, __):
        self.xlsx_file = _[0]
        self.root.clear_widgets()
        self.build()


if __name__ == "__main__":
    ScheduleApp().run()
