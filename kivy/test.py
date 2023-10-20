import kivy
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from datetime import datetime, timedelta
from openpyxl import load_workbook
from kivy.core.window import Window
import os


class ScheduleApp(App):
    def build(self):
        layout = BoxLayout(orientation="vertical", spacing=20, padding=20)

        # Get the current date
        end_date = datetime.now() + timedelta(days=6)       
        # Add a label for the current date with improved accessibility
        date_label = Label(
            text=end_date.strftime("%A, %d %B %Y"),
            color=(0, 0, 0, 1),  # Set text color to black for better contrast
            size_hint=(1, None),
            height=100,  # Increase the label's height for easier reading
            font_size=36,  # Use a larger font size for better readability
        )
        layout.add_widget(date_label)

        try:
            xlsx_file_path = self.locate_schedule_xlsx()
            if xlsx_file_path:
                wb = load_workbook(xlsx_file_path)
                sheet = wb.active

                # Create a ScrollView to display the schedule
                scrollview = ScrollView(size_hint=(1, 1))
                grid = GridLayout(cols=1, spacing=20, size_hint_y=None)

                data_found = False

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_date = row[0]
                    row_time = row[1]
                    row_discipline = row[2]
                    row_aud = row[3]
                    row_teacher = row[7]
                    row_column_G = row[6]  # Column G

                    # Check if the row date falls within the specified date range
                    if (
                        isinstance(row_date, datetime)
                        and row_date.date() == end_date.date()
                        and ("__________" in row_column_G or "2 подгруппа" in row_column_G)
                    ):
                        data_found = True

                        # Create a BoxLayout for each row of data with improved accessibility
                        row_layout = BoxLayout(orientation="vertical", spacing=10, padding=10)
                        
                        time_label = Label(text=row_time, font_size=24)
                        discipline_label = Label(text=row_discipline, font_size=24)
                        aud_label = Label(text=f"Audience: {row_aud}" if row_aud else "Audience", font_size=24)
                        teacher_label = Label(text=f"Teacher: {row_teacher}", font_size=24)
                        
                        # Add spacing between labels
                        row_layout.add_widget(time_label)
                        row_layout.add_widget(discipline_label)
                        row_layout.add_widget(aud_label)
                        row_layout.add_widget(teacher_label)

                        grid.add_widget(row_layout)

                scrollview.add_widget(grid)
                layout.add_widget(scrollview)

                if not data_found:
                    layout.add_widget(
                        Label(
                            text="[b]Nothing found for today[/b]",
                            size_hint=(1, None),
                            height=150,
                            markup=True,
                            font_size=28,  # Increase the font size for better visibility
                            color=(0, 0, 0, 1),  # Set text color to black for better contrast
                            halign="center",  # Center-align the text
                        )
                    )

        except Exception as e:
            error_popup = Popup(
                title="Error",
                content=Label(text=f"An error occurred: {str(e)}", font_size=20),
                size_hint=(None, None),
                size=(400, 200),
            )
            error_popup.open()
            print(str(e))

        return layout

    def locate_schedule_xlsx(self):
        documents_dir = os.path.join(os.path.expanduser("~"), "Documents")

        xlsx_path = os.path.join(documents_dir, "schedule.xlsx")
        print(xlsx_path)
        if os.path.exists(xlsx_path):
            return xlsx_path

        return None

if __name__ == "__main__":
    ScheduleApp().run()